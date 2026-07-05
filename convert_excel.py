"""Excel → i54 브랜드별 JSON 누적 변환 스크립트

브랜드별로 두 파일을 생성한다:
  {brand}.json      — 목록용 (detail_images 제외, 빠른 로드)
  {brand}.full.json — 상세용 (전체 필드)

Usage:
  python3 convert_excel.py data/2026-05-27.xlsx   # 새 엑셀 누적
  python3 convert_excel.py                         # data/ 목록에서 선택
  python3 convert_excel.py --no-thumbnails         # 썸네일 다운로드 생략
  python3 convert_excel.py --clean-thumbnails      # thumbnails/ 초기화 후 전체 재다운로드
  python3 convert_excel.py --duplicate always      # 중복 시 무조건 새 데이터로 갱신
  python3 convert_excel.py --duplicate never       # 중복 시 무조건 스킵 (기존 유지)
  python3 convert_excel.py --duplicate newer       # 중복 시 mfg_date 최신 데이터 유지 (기본값)
  python3 convert_excel.py --rebuild               # 기존 .json → listing/.full 분리 재생성
  python3 convert_excel.py --purge 2026-05-21      # 해당 날짜 이하 상품 전체 삭제
  python3 convert_excel.py --delete "러빈.브리즈줄팬츠" "슈크림.소다팝줄티"  # 특정 상품 삭제
"""
import sys
import openpyxl
import json
import re
import os
import shutil
import urllib.request
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor

THUMBNAILS_DIR = "thumbnails"

BRANDS_DIR = "public/data/i54/brands"
BRANDS_JSON = "public/data/brands.json"
SOLDOUT_JSON = "public/data/soldout.json"
SEARCH_INDEX_JSON = "public/data/search_index.json"
ITEMS_DIR = "data"

LISTING_FIELDS = {"id", "brand", "name", "price_sale", "thumbnail_url", "colors", "sizes", "mfg_date"}


def get_id_prefix(xlsx_path: str) -> str:
    """파일명에서 ID 접두어 추출.
    2026-05.xlsx             → '260500'
    2026-05-27.xlsx          → '260527'
    page_crawl_2026-06-30.xlsx → '260630'
    """
    stem = Path(xlsx_path).stem
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', stem)
    if m:
        return f"{m.group(1)[2:]}{m.group(2)}{m.group(3)}"
    m = re.search(r'(\d{4})-(\d{2})', stem)
    if m:
        return f"{m.group(1)[2:]}{m.group(2)}00"
    return re.sub(r"\D", "", stem)[:6]


def load_brand(brand: str) -> list:
    """.full.json 우선 로드, 없으면 .json (마이그레이션 대응)."""
    for suffix in (".full.json", ".json"):
        path = os.path.join(BRANDS_DIR, f"{brand}{suffix}")
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                return json.load(f)
    return []


def save_brand(brand: str, products: list) -> None:
    """전체 데이터는 .full.json, 목록용은 .json 으로 분리 저장."""
    full_path = os.path.join(BRANDS_DIR, f"{brand}.full.json")
    with open(full_path, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)

    listing = [{k: v for k, v in p.items() if k in LISTING_FIELDS} for p in products]
    listing_path = os.path.join(BRANDS_DIR, f"{brand}.json")
    with open(listing_path, "w", encoding="utf-8") as f:
        json.dump(listing, f, ensure_ascii=False, indent=2)


def parse_options(opt_name, opt_val, opt_price):
    colors, sizes = [], []
    if not opt_name or not opt_val:
        return colors, sizes

    names = [n.strip() for n in re.split(r"\r?\n", str(opt_name)) if n.strip()]
    vals_parts = re.split(r"\r?\n", str(opt_val))
    price_parts = re.split(r"\r?\n", str(opt_price)) if opt_price else []

    for i, name in enumerate(names):
        raw_vals = vals_parts[i] if i < len(vals_parts) else ""
        vals = [v.strip() for v in raw_vals.split(",") if v.strip()]

        prices = []
        if i < len(price_parts) and price_parts[i].strip():
            try:
                prices = [int(float(p.strip())) for p in price_parts[i].split(",") if p.strip()]
            except ValueError:
                pass
        while len(prices) < len(vals):
            prices.append(0)

        options = [{"name": v, "add_price": prices[j]} for j, v in enumerate(vals)]
        if "색상" in name:
            colors = options
        elif "사이즈" in name:
            sizes = options

    return colors, sizes


def extract_detail_images(html) -> list:
    if not html:
        return []
    return re.findall(r"""<img[^>]+src=['"]([^'"]+)['"]""", str(html))


def download_thumbnails(new_by_brand: dict, clean: bool = False) -> None:
    """이번 엑셀에 등장한 모든 상품(중복 스킵 포함)의 썸네일을 로컬에 저장.
    파일명: {brand}_{상품명}{확장자}
    저장 위치: thumbnails/
    clean=True 이면 기존 디렉토리를 삭제 후 전체 재다운로드, False 이면 이미 있는 파일은 스킵.
    """
    out_dir = Path(THUMBNAILS_DIR)
    if clean and out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    tasks = []
    skipped = 0
    for brand, products in new_by_brand.items():
        for p in products:
            url = p.get("thumbnail_url")
            if not url:
                continue
            safe_name = re.sub(r'[\\/:*?"<>|.]', "_", p["name"])
            ext = os.path.splitext(url.split("?")[0])[1] or ".jpg"
            dest = out_dir / f"{safe_name}{ext}"
            if not clean and dest.exists():
                skipped += 1
                continue
            tasks.append((url, dest))

    def fetch(task):
        url, dest = task
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as resp, open(dest, "wb") as f:
                f.write(resp.read())
            return None
        except Exception as e:
            return f"{dest.name}: {e}"

    skip_msg = f", {skipped}개 이미 존재(스킵)" if skipped else ""
    print(f"썸네일 다운로드 중... ({len(tasks)}개{skip_msg} → {out_dir})")
    errors = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        for result in executor.map(fetch, tasks):
            if result:
                errors.append(result)

    print(f"썸네일 저장 완료: {len(tasks) - len(errors)}개")
    if errors:
        print(f"  실패 {len(errors)}개:")
        for e in errors[:10]:
            print(f"    {e}")


def pick_xlsx(items_dir: str) -> str:
    files = sorted(Path(items_dir).glob("*.xlsx"))
    if not files:
        print(f"오류: {items_dir} 에 xlsx 파일이 없습니다.")
        sys.exit(1)
    if len(files) == 1:
        print(f"자동 선택: {files[0]}")
        return str(files[0])
    print("변환할 파일을 선택하세요:")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {f.name}")
    while True:
        try:
            n = int(input("번호 입력: "))
            if 1 <= n <= len(files):
                return str(files[n - 1])
        except (ValueError, KeyboardInterrupt):
            sys.exit(0)


def rebuild_brands_json() -> None:
    """listing .json 파일들을 읽어 brands.json 재생성."""
    brands_json = []
    for fname in sorted(os.listdir(BRANDS_DIR)):
        if fname.endswith(".full.json") or not fname.endswith(".json"):
            continue
        brand = fname[:-5]
        with open(os.path.join(BRANDS_DIR, fname), encoding="utf-8") as f:
            products = json.load(f)
        if not products:
            continue
        sorted_products = sorted(reversed(products), key=lambda p: p.get("mfg_date", ""), reverse=True)
        thumbs = [p["thumbnail_url"] for p in sorted_products if p.get("thumbnail_url")][:3]
        dates = [p["mfg_date"] for p in products if p.get("mfg_date")]
        brands_json.append({
            "id": brand,
            "name": brand,
            "total": len(products),
            "preview_thumbnails": thumbs,
            "latest_mfg_date": max(dates) if dates else "",
        })
    with open(BRANDS_JSON, "w", encoding="utf-8") as f:
        json.dump(brands_json, f, ensure_ascii=False, indent=2)
    print(f"brands.json 갱신: {len(brands_json)}개 브랜드")


def rebuild_search_index() -> None:
    """listing .json 파일들을 모아 검색용 search_index.json 생성."""
    index = []
    for fname in sorted(os.listdir(BRANDS_DIR)):
        if fname.endswith(".full.json") or not fname.endswith(".json"):
            continue
        brand = fname[:-5]
        with open(os.path.join(BRANDS_DIR, fname), encoding="utf-8") as f:
            products = json.load(f)
        for p in reversed(products):
            index.append({
                "id": p["id"],
                "brand": brand,
                "name": p.get("name", ""),
                "price_sale": p.get("price_sale", 0),
                "thumbnail_url": p.get("thumbnail_url", ""),
                "mfg_date": p.get("mfg_date", ""),
                "colors": p.get("colors", []),
                "sizes": p.get("sizes", []),
            })
    with open(SEARCH_INDEX_JSON, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)
    print(f"search_index.json 갱신: {len(index)}개 상품")


def clean_soldout(delete_ids: set) -> None:
    """삭제된 상품 ID를 soldout.json에서 제거."""
    if not delete_ids or not os.path.exists(SOLDOUT_JSON):
        return
    with open(SOLDOUT_JSON, encoding="utf-8") as f:
        soldout = json.load(f)
    cleaned = [sid for sid in soldout if sid not in delete_ids]
    if len(cleaned) != len(soldout):
        with open(SOLDOUT_JSON, "w", encoding="utf-8") as f:
            json.dump(cleaned, f, ensure_ascii=False, indent=2)
        print(f"soldout.json: {len(soldout) - len(cleaned)}개 제거")


def rebuild_split() -> None:
    """기존 .json (full data) → listing .json + .full.json 으로 분리 재생성."""
    print("기존 브랜드 파일 분리 재생성 중...")
    count = 0
    for fname in sorted(os.listdir(BRANDS_DIR)):
        if fname.endswith(".full.json") or not fname.endswith(".json"):
            continue
        brand = fname[:-5]
        with open(os.path.join(BRANDS_DIR, fname), encoding="utf-8") as f:
            products = json.load(f)
        save_brand(brand, products)
        print(f"  {brand}: {len(products)}개")
        count += 1
    rebuild_brands_json()
    rebuild_search_index()
    print(f"\n완료: {count}개 브랜드 분리 완료")


def delete_items(names: list) -> None:
    """지정한 name 값을 가진 상품을 모든 브랜드 파일에서 제거.
    python3 convert_excel.py --delete "러빈.브리즈줄팬츠" "슈크림.소다팝줄티"
    """
    target = set(names)
    removed_total = 0
    not_found = set(target)
    delete_ids = set()

    for fname in sorted(os.listdir(BRANDS_DIR)):
        if not fname.endswith(".full.json"):
            continue
        brand = fname[:-10]
        path = os.path.join(BRANDS_DIR, fname)
        with open(path, encoding="utf-8") as f:
            products = json.load(f)

        kept = [p for p in products if p.get("name") not in target]
        removed = len(products) - len(kept)

        if removed > 0:
            save_brand(brand, kept)
            found = [p for p in products if p.get("name") in target]
            for p in found:
                not_found.discard(p["name"])
                delete_ids.add(p["id"])
            print(f"  {brand}: -{removed}개 제거")
            removed_total += removed

    if not_found:
        for n in sorted(not_found):
            print(f"  경고: '{n}' 를 찾지 못했습니다.")

    if removed_total:
        rebuild_brands_json()
        rebuild_search_index()
        clean_soldout(delete_ids)
    print(f"\n완료: {removed_total}개 제거")


def purge_before(cutoff_date: str) -> None:
    """cutoff_date 이하 mfg_date 를 가진 상품을 모든 브랜드 파일에서 제거.
    python3 convert_excel.py --purge 2026-05-21
    """
    import re as _re
    if not _re.match(r"\d{4}-\d{2}-\d{2}", cutoff_date):
        print(f"오류: 날짜 형식이 올바르지 않습니다 (예: 2026-05-21)")
        sys.exit(1)

    print(f"{cutoff_date} 이하 상품 제거 중...")
    removed_total = kept_total = 0
    affected = 0
    delete_ids = set()

    for fname in sorted(os.listdir(BRANDS_DIR)):
        if not fname.endswith(".full.json"):
            continue
        brand = fname[:-10]
        path = os.path.join(BRANDS_DIR, fname)
        with open(path, encoding="utf-8") as f:
            products = json.load(f)

        kept = [p for p in products if p.get("mfg_date", "") > cutoff_date]
        removed_products = [p for p in products if p.get("mfg_date", "") <= cutoff_date]
        removed = len(removed_products)

        if removed > 0:
            save_brand(brand, kept)
            delete_ids.update(p["id"] for p in removed_products)
            print(f"  {brand}: -{removed}개 제거 (잔여 {len(kept)}개)")
            removed_total += removed
            affected += 1
        kept_total += len(kept)

    rebuild_brands_json()
    rebuild_search_index()
    clean_soldout(delete_ids)
    print(f"\n완료: {affected}개 브랜드에서 {removed_total}개 제거, {kept_total}개 유지")


def main():
    args = sys.argv[1:]
    no_thumbnails = "--no-thumbnails" in args
    if no_thumbnails:
        args.remove("--no-thumbnails")

    clean_thumbnails = "--clean-thumbnails" in args
    if clean_thumbnails:
        args.remove("--clean-thumbnails")

    if no_thumbnails and clean_thumbnails:
        print("오류: --no-thumbnails 와 --clean-thumbnails 는 함께 사용할 수 없습니다.")
        sys.exit(1)

    duplicate_mode = "newer"
    if "--duplicate" in args:
        idx = args.index("--duplicate")
        if idx + 1 >= len(args):
            print("오류: --duplicate 다음에 모드를 지정하세요 (always|never|newer)")
            sys.exit(1)
        duplicate_mode = args[idx + 1]
        if duplicate_mode not in ("always", "never", "newer"):
            print(f"오류: --duplicate 모드는 always, never, newer 중 하나여야 합니다.")
            sys.exit(1)
        args = args[:idx] + args[idx + 2:]

    if len(args) > 0 and args[0] == "--rebuild":
        rebuild_split()
        return

    if len(args) > 1 and args[0] == "--purge":
        purge_before(args[1])
        return

    if len(args) > 1 and args[0] == "--delete":
        delete_items(args[1:])
        return

    xlsx_path = args[0] if len(args) > 0 else pick_xlsx(ITEMS_DIR)

    if not os.path.exists(xlsx_path):
        print(f"오류: {xlsx_path} 파일을 찾을 수 없습니다.")
        sys.exit(1)

    id_prefix = get_id_prefix(xlsx_path)
    print(f"처리 중: {xlsx_path}  (ID 접두어: {id_prefix})")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    new_by_brand: dict[str, list] = defaultdict(list)
    for row_idx in range(3, ws.max_row + 1):
        row = ws[row_idx]
        brand = (row[25].value or "").strip()
        name = (row[2].value or "").strip()
        if not brand or not name:
            continue

        price_str = str(row[4].value or "0").strip()
        try:
            price_sale = int(float(price_str))
        except ValueError:
            price_sale = 0

        colors, sizes = parse_options(row[13].value, row[14].value, row[15].value)
        thumbnail_url = (row[22].value or "").strip()
        detail_images = extract_detail_images(row[24].value)

        mfg_date = str(row[27].value or "").strip()[:10]

        new_by_brand[brand].append({
            "id": f"{id_prefix}{row_idx - 2:04d}",
            "brand": brand,
            "name": f"{brand}.{name}",
            "price_sale": price_sale,
            "thumbnail_url": thumbnail_url,
            "detail_images": detail_images,
            "colors": colors,
            "sizes": sizes,
            "mfg_date": mfg_date,
        })

    if no_thumbnails:
        print("썸네일 다운로드 스킵 (--no-thumbnails)")
    else:
        download_thumbnails(new_by_brand, clean=clean_thumbnails)

    os.makedirs(BRANDS_DIR, exist_ok=True)
    added_total = updated_total = skipped_total = 0

    for brand, new_products in new_by_brand.items():
        existing = load_brand(brand)
        result = list(existing)

        def product_key(p):
            colors = tuple(c["name"] for c in p.get("colors", []))
            sizes = tuple(s["name"] for s in p.get("sizes", []))
            return (p["name"], colors, sizes)

        name_to_index = {product_key(p): i for i, p in enumerate(result)}

        # 파일 내 중복은 mfg_date가 더 최신인 쪽만 남긴다
        latest_by_key = {}
        for p in new_products:
            key = product_key(p)
            prev = latest_by_key.get(key)
            if prev is None or p.get("mfg_date", "") > prev.get("mfg_date", ""):
                latest_by_key[key] = p

        added = updated = skipped = 0
        for key, p in latest_by_key.items():
            if key in name_to_index:
                idx = name_to_index[key]
                old = result[idx]
                if duplicate_mode == "always":
                    result[idx] = p
                    updated += 1
                elif duplicate_mode == "never":
                    skipped += 1
                else:  # newer
                    if p.get("mfg_date", "") > old.get("mfg_date", ""):
                        result[idx] = p
                        updated += 1
                    else:
                        skipped += 1
            else:
                result.append(p)
                name_to_index[key] = len(result) - 1
                added += 1

        if added or updated:
            save_brand(brand, result)

        added_total += added
        updated_total += updated
        skipped_total += skipped
        status = f"+{added} 추가"
        if updated:
            status += f", {updated} 갱신"
        if skipped:
            skip_reason = "스킵(강제)" if duplicate_mode == "never" else "스킵(이전 데이터)"
            status += f", {skipped} {skip_reason}"
        print(f"  {brand}: {status}  (총 {len(result)}개)")

    rebuild_brands_json()
    rebuild_search_index()
    print(f"\n완료: +{added_total}개 추가, {updated_total}개 갱신, {skipped_total}개 스킵")


if __name__ == "__main__":
    main()
