"""i54.co.kr 브랜드별 상품 크롤러 → my-shushu convert_excel.py 입력용 엑셀 생성

Usage:
  python3 crawl_brand.py <브랜드명> [--since YYYY-MM-DD]

  --since YYYY-MM-DD : 해당 날짜 이후(포함)에 제조된 상품만 수집

결과: my-shushu/data/<브랜드명>_<오늘날짜>.xlsx (없으면 새로 생성, 있으면 이어서 추가)
이후 my-shushu 디렉터리에서 `python3 convert_excel.py <파일명>.xlsx` 실행하면
public/data/i54/brands/<브랜드명>.json 으로 반영된다.
"""
import sys
import os
import re
import json
import time
from datetime import date
from pathlib import Path

import requests
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ENV_PATH = ROOT / ".env"
OUT_DIR = ROOT / "my-shushu" / "data"

BASE_URL = "https://www.i54.co.kr"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
}

REQUEST_DELAY = 1.0


def load_env(path):
    env = {}
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip()
    return env


def login(session, member_id, member_passwd):
    session.get(f"{BASE_URL}/member/login.html", headers=HEADERS)
    r = session.post(
        f"{BASE_URL}/exec/front/Member/login/",
        data={"member_id": member_id, "member_passwd": member_passwd},
        headers=HEADERS,
    )
    r.raise_for_status()
    home = session.get(BASE_URL, headers=HEADERS)
    if "로그아웃" not in home.text:
        raise RuntimeError("로그인 실패: 아이디/비밀번호를 확인해주세요.")


def search_product_nos(session, brand):
    """검색 결과에서 '브랜드.' 로 시작하는 상품의 product_no 목록을 수집."""
    product_nos = []
    page = 1
    while True:
        r = session.get(
            f"{BASE_URL}/product/search.html",
            params={"keyword": brand, "page": page},
            headers=HEADERS,
        )
        r.raise_for_status()
        html = r.text
        items = re.findall(r'<li id="anchorBoxId_(\d+)".*?</li>', html, re.S)
        if not items:
            break

        blocks = re.findall(r'<li id="anchorBoxId_\d+".*?</li>', html, re.S)
        for pno, block in zip(items, blocks):
            m = re.search(r'<strong class="name">.*?</span> <span[^>]*>(.*?)</span></a>', block, re.S)
            title = re.sub(r"<[^>]+>", "", m.group(1)).strip() if m else ""
            if title.startswith(f"{brand}."):
                product_nos.append(pno)

        page += 1
        time.sleep(REQUEST_DELAY)

    return product_nos


def parse_detail(session, product_no, brand):
    r = session.get(
        f"{BASE_URL}/product/detail.html",
        params={"product_no": product_no},
        headers=HEADERS,
    )
    r.raise_for_status()
    html = r.text

    m = re.search(r'<th scope="row"><span[^>]*>상품명</span></th>\s*<td><span[^>]*>(.*?)</span>', html, re.S)
    raw_name = m.group(1) if m else ""
    name = re.split(r"<br\s*/?>", raw_name)[0].strip()
    name = re.sub(r"<[^>]+>", "", name).strip()

    m = re.search(r'id="span_product_price_text">([\d,]+)원', html)
    sale_price = int(m.group(1).replace(",", "")) if m else 0

    m = re.search(r'id="span_product_price_custom"><strike>([\d,]+)원', html)
    price_sale = int(m.group(1).replace(",", "")) if m else sale_price

    m = re.search(r'<meta property="og:image" content="([^"]+)"', html)
    thumbnail_url = m.group(1) if m else ""

    detail_images = re.findall(r'ec-data-src="([^"]+)"', html)
    detail_images = [u if u.startswith("http") else f"https:{u}" for u in detail_images]
    detail_images = [u for u in detail_images if not re.search(r"(washingtip|kc)\.jpg", u, re.I)]

    m = re.search(r'<th scope="row"><span[^>]*>제조일자</span></th>\s*<td><span[^>]*>(.*?)</span>', html, re.S)
    mfg_date = re.sub(r"<[^>]+>", "", m.group(1)).strip()[:10] if m else ""

    colors, sizes = parse_options(html, sale_price)

    return {
        "id": product_no,
        "brand": brand,
        "name": f"{brand}.{name}" if not name.startswith(brand) else name,
        "price_sale": price_sale,
        "thumbnail_url": thumbnail_url,
        "detail_images": detail_images,
        "colors": colors,
        "sizes": sizes,
        "mfg_date": mfg_date,
    }


def parse_options(html, price_sale):
    m = re.search(r"var option_stock_data = '(.*?)';", html, re.S)
    if not m:
        return [], []
    try:
        # 페이지 내 JS 문자열 리터럴 안에 JSON이 이중 escape되어 있음
        raw = m.group(1).replace('\\"', '"')
        raw = raw.encode("utf-8").decode("unicode_escape").encode("latin-1").decode("utf-8")
        data = json.loads(raw)
    except (json.JSONDecodeError, UnicodeError):
        return [], []

    dims = {}  # dim_name -> {value: add_price}
    for combo in data.values():
        names = combo.get("option_name_original") or []
        values = combo.get("option_value_orginal") or []
        add_price = (combo.get("option_price") or 0) - price_sale
        for dim_name, value in zip(names, values):
            bucket = dims.setdefault(dim_name, {})
            if value not in bucket:
                bucket[value] = add_price
            else:
                bucket[value] = min(bucket[value], add_price)

    colors = [{"name": clean_option_name(v), "add_price": p} for v, p in dims.get("색상", {}).items()]
    sizes = [{"name": clean_option_name(v), "add_price": p} for v, p in dims.get("사이즈", {}).items()]
    return colors, sizes


def clean_option_name(value):
    """옵션값에 포함된 '(+3000)' 같은 추가금 표기를 제거."""
    return re.sub(r"\(\s*[+-]?\s*\d+\s*\)", "", value).strip()


def write_xlsx(products, out_path):
    if out_path.exists():
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "상품 기본정보"
        headers = {3: "상품명", 5: "판매가", 14: "옵션명", 15: "옵션값", 16: "옵션가",
                   23: "대표이미지", 25: "상세설명", 26: "브랜드", 28: "제조일자"}
        for col, label in headers.items():
            ws.cell(row=2, column=col, value=label)
        start_row = 3

    for i, p in enumerate(products):
        row = start_row + i

        opt_names, opt_vals, opt_prices = [], [], []
        if p["colors"]:
            opt_names.append("색상")
            opt_vals.append(",".join(c["name"] for c in p["colors"]))
            opt_prices.append(",".join(str(c["add_price"]) for c in p["colors"]))
        if p["sizes"]:
            opt_names.append("사이즈")
            opt_vals.append(",".join(s["name"] for s in p["sizes"]))
            opt_prices.append(",".join(str(s["add_price"]) for s in p["sizes"]))

        ws.cell(row=row, column=3, value=p["name"].split(".", 1)[-1])
        ws.cell(row=row, column=5, value=p["price_sale"])
        if opt_names:
            ws.cell(row=row, column=14, value="\n".join(opt_names))
            ws.cell(row=row, column=15, value="\n".join(opt_vals))
            ws.cell(row=row, column=16, value="\n".join(opt_prices))
        ws.cell(row=row, column=23, value=p["thumbnail_url"])
        ws.cell(row=row, column=25, value="".join(f"<img src='{u}'>" for u in p["detail_images"]))
        ws.cell(row=row, column=26, value=p["brand"])
        ws.cell(row=row, column=28, value=p["mfg_date"])

    wb.save(out_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 crawl_brand.py <브랜드명> [--since YYYY-MM-DD]")
        sys.exit(1)
    brand = sys.argv[1]

    since = None
    if "--since" in sys.argv:
        since = sys.argv[sys.argv.index("--since") + 1]

    env = load_env(ENV_PATH)
    session = requests.Session()
    login(session, env["I54_ID"], env["I54_PW"])
    print(f"로그인 성공. '{brand}' 검색 중...")

    product_nos = search_product_nos(session, brand)
    print(f"'{brand}.'로 시작하는 상품 {len(product_nos)}개 발견")

    products = []
    skipped = 0
    for i, pno in enumerate(product_nos, 1):
        p = parse_detail(session, pno, brand)
        if since and p["mfg_date"] and p["mfg_date"] < since:
            skipped += 1
            print(f"  [{i}/{len(product_nos)}] {p['name']} - 제조일자 {p['mfg_date']} (제외)")
            time.sleep(REQUEST_DELAY)
            continue
        products.append(p)
        print(f"  [{i}/{len(product_nos)}] {p['name']} - {p['price_sale']}원 (제조일자 {p['mfg_date']})")
        time.sleep(REQUEST_DELAY)

    if since:
        print(f"\n제조일자 {since} 이전 상품 {skipped}개 제외")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"{brand}_{date.today().isoformat()}.xlsx"
    write_xlsx(list(reversed(products)), out_path)
    print(f"\n완료: {out_path} 에 {len(products)}개 상품 저장")


if __name__ == "__main__":
    main()
